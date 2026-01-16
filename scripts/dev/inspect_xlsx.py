from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter


def cell_text(cell: Cell) -> str:
    v = cell.value
    if v is None:
        return ""
    s = str(v)
    return " ".join(s.split())


def iter_cells(ws, max_row: int, max_col: int):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            yield ws.cell(r, c)


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect an XLSX and dump a human-readable summary.")
    parser.add_argument(
        "--xlsx",
        default="docs/Quarantine/UNCLASSIFIED/Deals/Client Pitch Decks - Whitepapers/Financials - WebMax Valuation and Allocation of funds.xlsx",
        help="Path to an .xlsx file (relative to repo root).",
    )
    parser.add_argument(
        "--out",
        default="artifacts/xlsx_inspection_webmax.md",
        help="Output markdown path (relative to repo root).",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)

    if not xlsx_path.exists():
        raise SystemExit(f"XLSX not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=False)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    md: list[str] = []
    md.append(f"# XLSX Inspection Report\n")
    md.append(f"- Workbook: `{xlsx_path}`\n")
    md.append(f"- Size: `{xlsx_path.stat().st_size}` bytes\n")
    md.append(f"- Sheets: {', '.join([f'`{s}`' for s in wb.sheetnames])}\n")

    keywords = [
        "revenue",
        "sales",
        "cogs",
        "gross",
        "margin",
        "opex",
        "operating",
        "ebitda",
        "net income",
        "cash",
        "burn",
        "runway",
        "valuation",
        "cap table",
        "allocation",
        "use of funds",
    ]

    for name in wb.sheetnames:
        ws = wb[name]
        dim = ws.calculate_dimension()
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        tables = list(getattr(ws, "tables", {}).keys())

        md.append("---\n")
        md.append(f"## Sheet: `{name}`\n")
        md.append(f"- Dimension: `{dim}`\n")
        md.append(f"- Max row/col: `{max_row}` / `{max_col}`\n")
        if tables:
            md.append(f"- Excel tables: {', '.join([f'`{t}`' for t in tables])}\n")

        # Quick content scan
        scan_rows = min(max_row, 200)
        scan_cols = min(max_col, 30)
        non_empty = 0
        hits: list[tuple[str, str]] = []

        for cell in iter_cells(ws, scan_rows, scan_cols):
            txt = cell_text(cell)
            if txt:
                non_empty += 1
                low = txt.lower()
                for kw in keywords:
                    if kw in low:
                        addr = f"{get_column_letter(cell.column)}{cell.row}"
                        hits.append((addr, txt))
                        break

        md.append(f"- Non-empty in `A1..{get_column_letter(scan_cols)}{scan_rows}`: `{non_empty}`\n")
        if hits:
            md.append("**Keyword hits (first 30):**\n")
            for addr, txt in hits[:30]:
                md.append(f"- `{addr}`: {txt}\n")
            if len(hits) > 30:
                md.append(f"- … ({len(hits) - 30} more)\n")

        # Small top-left preview (only rows that have any content)
        preview_rows = min(max_row, 35)
        preview_cols = min(max_col, 12)
        if preview_rows and preview_cols:
            md.append("\n**Preview (top-left, up to 35 rows x 12 cols):**\n")
            md.append("```text\n")
            for r in range(1, preview_rows + 1):
                row_vals = []
                for c in range(1, preview_cols + 1):
                    s = cell_text(ws.cell(r, c))
                    if len(s) > 40:
                        s = s[:37] + "..."
                    row_vals.append(s)
                if any(row_vals):
                    md.append(f"{r:>3} | " + " | ".join(row_vals) + "\n")
            md.append("```\n")

    out_path.write_text("".join(md), encoding="utf-8")
    print(f"Wrote report: {out_path}")


if __name__ == "__main__":
    main()
